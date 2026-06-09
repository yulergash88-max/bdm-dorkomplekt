import io
from datetime import datetime

from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bot.keyboards.common import ROLE_LABELS
from bot.utils.formatting import STATUS_LABELS


def _make_workbook(header: list[str], rows: list[list]) -> Workbook:
    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(header)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row in rows:
        ws.append(row)

    # Auto-fit column widths
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    return wb


def _wb_to_file(wb: Workbook, filename: str) -> BufferedInputFile:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return BufferedInputFile(buf.read(), filename=filename)


def deliveries_to_csv(deliveries: list[dict]) -> BufferedInputFile:
    return _deliveries_xlsx(deliveries, "yetkazib_berishlar.xlsx")


def deliveries_to_csv_by_date(deliveries: list[dict], date_from: str, date_to: str) -> BufferedInputFile:
    filename = f"hisobot_{date_from.replace('.', '')}_{date_to.replace('.', '')}.xlsx"
    return _deliveries_xlsx(deliveries, filename)


def _deliveries_xlsx(deliveries: list[dict], filename: str) -> BufferedInputFile:
    header = [
        "№",
        "Товар",
        "Машина рақами",
        "Сана",
        "Ҳолат",
        "Юборилди (м³)",
        "Тортилди (т)",
        "Коэффициент",
        "Харидор куби (м³)",
        "Фарқ (м³)",
    ]
    rows = []
    for d in deliveries:
        raw_date = d.get("created_at") or ""
        try:
            date_str = datetime.fromisoformat(raw_date[:19]).strftime("%d.%m.%Y %H:%M")
        except (ValueError, TypeError):
            date_str = raw_date[:10]
        rows.append([
            d["id"],
            d.get("product_name") or "",
            d.get("car_number") or "",
            date_str,
            STATUS_LABELS.get(d["status"], d["status"]),
            d.get("supplier_kub") or "",
            d.get("buyer_tonnage") or "",
            d.get("lab_coefficient") or "",
            d.get("buyer_kub") or "",
            d.get("kub_difference") or "",
        ])
    return _wb_to_file(_make_workbook(header, rows), filename)


def users_to_csv(users: list[dict]) -> BufferedInputFile:
    header = ["Telegram ID", "Исм-фамилия", "Роли", "Телефон", "Тасдиқланган", "Блокланган", "Яратилган сана"]
    rows = [
        [
            user["telegram_id"],
            user["full_name"],
            ROLE_LABELS[user["role"]],
            user["phone"] or "",
            "ҳа" if user["is_approved"] else "йўқ",
            "ҳа" if user["is_blocked"] else "йўқ",
            user["created_at"],
        ]
        for user in users
    ]
    return _wb_to_file(_make_workbook(header, rows), "foydalanuvchilar.xlsx")
